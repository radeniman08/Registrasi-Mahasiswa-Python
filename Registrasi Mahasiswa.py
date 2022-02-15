# import modul openpyxl dan tkinter
import openpyxl
from tkinter import *

# mendeklarasikan variabel wb dan sheet

# membuka file excel yang ada
wb = openpyxl.Workbook()

# membuat objek sheet
ws = wb.active


def excel():
    # ubah ukuran lebar kolom diexcel spreadsheet
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 50

    # tulis data yang diberikan ke spreadsheet excel dilokasi tertentu
    ws.cell( row=1 , column=1 ).value = "NAMA"
    ws.cell( row=1 , column=2 ).value = "JURUSAN"
    ws.cell( row=1 , column=3 ).value = "SEMESTER"
    ws.cell( row=1 , column=4 ).value = "NIM"
    ws.cell( row=1 , column=5 ).value = "NOMER TELPON"
    ws.cell( row=1 , column=6 ).value = "EMAIL"
    ws.cell( row=1 , column=7 ).value = "ALAMAT"


# Berfungsi untuk mengatur fokus (kursor)
def fokus1(event):
    # atur fokus pada kotak course_field
    course_field.focus_set()


# Berfungsi untuk mengatur fokus
def fokus2(event):
    # atur fokus pada kotak sem_field
    sem_field.focus_set()


# Berfungsi untuk mengatur fokus
def fokus3(event):
    # atur fokus pada kotak form_no_field
    form_no_field.focus_set()


# Berfungsi untuk mengatur fokus
def fokus4(event):
    # atur fokus pada kotak contact_no_field
    contact_no_field.focus_set()


# Berfungsi untuk mengatur fokus
def fokus5(event):
    # atur fokus pada kotak email_id_field
    email_id_field.focus_set()


# Berfungsi untuk mengatur fokus
def fokus6(event):
    # atur fokus pada kotak address_field
    address_field.focus_set()


# Fungsi untuk membersihkan isi kotak entri teks
def clear():
    # kosongkan isi kotak entri teks
    name_field.delete( 0 , END )
    course_field.delete( 0 , END )
    sem_field.delete( 0 , END )
    form_no_field.delete( 0 , END )
    contact_no_field.delete( 0 , END )
    email_id_field.delete( 0 , END )
    address_field.delete( 0 , END )


# Berfungsi untuk mengambil data dari GUI dan tulis ke file excel
def insert():
    # jika pengguna tidak mengisi entri apa pun, lalu cetak "input kosong"
    if (name_field.get() == "" and
            course_field.get() == "" and
            sem_field.get() == "" and
            form_no_field.get() == "" and
            contact_no_field.get() == "" and
            email_id_field.get() == "" and
            address_field.get() == ""):

        print( "INPUTAN KOSONG" )

    else:

        # menugaskan baris maks dan kolom maks nilai hingga data mana yang ditulis dalam lembar excel ke variabel
        current_row = ws.max_row
        current_column = ws.max_column

        # get method mengembalikan teks saat ini
        # sebagai string tempat kita menulis
        # excel spreadsheet di lokasi tertentu
        ws.cell( row=current_row + 1 , column=1 ).value = name_field.get()
        ws.cell( row=current_row + 1 , column=2 ).value = course_field.get()
        ws.cell( row=current_row + 1 , column=3 ).value = sem_field.get()
        ws.cell( row=current_row + 1 , column=4 ).value = form_no_field.get()
        ws.cell( row=current_row + 1 , column=5 ).value = contact_no_field.get()
        ws.cell( row=current_row + 1 , column=6 ).value = email_id_field.get()
        ws.cell( row=current_row + 1 , column=7 ).value = address_field.get()

        # simpan file tersebut
        wb.save( 'C:\\Users\\R.Iman\\Desktop\\excel.xlsx' )

        # atur fokus pada kotak name_field
        name_field.focus_set()

        # panggil fungsi clear ()
        clear()

    # Driver kode


if __name__ == "__main__":
    # buat jendela GUI
    root = Tk()

    # mengatur warna latar belakang jendela GUI
    root.configure( background='light green' )

    # mengatur judul jendela GUI
    root.title( "MENU REGISTRASI MAHASISWA" )

    # atur konfigurasi jendela GUI
    root.geometry( "500x300" )

    excel()

    # buat label Form
    heading = Label( root , text="REGISTRASI MAHASISWA" , bg="light green" )

    # buat label Nama
    name = Label( root , text="NAMA" , bg="light green" )

    # buat label Jurusan
    course = Label( root , text="JURUSAN" , bg="light green" )

    # buat label Semester
    sem = Label( root , text="SEMESTER" , bg="light green" )

    # buat label Nim
    form_no = Label( root , text="NIM" , bg="light green" )

    # buat label Nomer Telpon
    contact_no = Label( root , text="NOMER TELPON" , bg="light green" )

    # buat label Email
    email_id = Label( root , text="EMAIL" , bg="light green" )

    # buat label Alamat
    address = Label( root , text="ALAMAT" , bg="light green" )

    # Metode grid digunakan untuk menempatkan Widget di posisi masing-masing dalam tabel seperti struktur.
    heading.grid( row=0 , column=1 )
    name.grid( row=1 , column=0 )
    course.grid( row=2 , column=0 )
    sem.grid( row=3 , column=0 )
    form_no.grid( row=4 , column=0 )
    contact_no.grid( row=5 , column=0 )
    email_id.grid( row=6 , column=0 )
    address.grid( row=7 , column=0 )

    # buat kotak entri teks untuk mengetik informasi
    name_field = Entry( root )
    course_field = Entry( root )
    sem_field = Entry( root )
    form_no_field = Entry( root )
    contact_no_field = Entry( root )
    email_id_field = Entry( root )
    address_field = Entry( root )

    # Metode bind dari widget digunakan untuk mengikat fungsi dengan acara setiap kali tombol enter ditekan lalu panggil fungsi fokus1
    name_field.bind( "<Return>" , fokus1 )

    # setiap kali tombol enter ditekan
    # lalu panggil fungsi fokus2
    course_field.bind( "<Return>" , fokus2 )

    # setiap kali tombol enter ditekan
    # lalu panggil fungsi fokus3
    sem_field.bind( "<Return>" , fokus3 )

    # setiap kali tombol enter ditekan
    # lalu panggil fungsi fokus4
    form_no_field.bind( "<Return>" , fokus4 )

    # setiap kali tombol enter ditekan
    # lalu panggil fungsi fokus5
    contact_no_field.bind( "<Return>" , fokus5 )

    # setiap kali tombol enter ditekan
    # lalu panggil fungsi fokus6
    email_id_field.bind( "<Return>" , fokus6 )

    # Metode grid digunakan untuk menempatkan Widget di posisi masing-masing dalam tabel seperti struktur.
    name_field.grid( row=1 , column=1 , ipadx="100" )
    course_field.grid( row=2 , column=1 , ipadx="100" )
    sem_field.grid( row=3 , column=1 , ipadx="100" )
    form_no_field.grid( row=4 , column=1 , ipadx="100" )
    contact_no_field.grid( row=5 , column=1 , ipadx="100" )
    email_id_field.grid( row=6 , column=1 , ipadx="100" )
    address_field.grid( row=7 , column=1 , ipadx="100" )

    # memanggil fungsi excel
    excel()

    # buat Tombol Kirim dan tempatkan ke jendela root
    submit = Button( root , text="Submit" , fg="Black" ,
                     bg="Red" , command=insert )
    submit.grid( row=8 , column=1 )

    # mulai GUI
    root.mainloop()